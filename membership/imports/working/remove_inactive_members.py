#!/usr/bin/env python3
"""
Remove inactive members from the canonical member list.

This script:
1. Reads the inactive member mailing list (654 emails)
2. Reads the canonical member list (2,562 records)
3. Removes any members who appear in the inactive list
4. Outputs cleaned canonical list
"""

import csv
import os
from openpyxl import load_workbook

WORKING_DIR = os.path.dirname(os.path.abspath(__file__))
INACTIVE_XLSX = os.path.join(WORKING_DIR, 'Mailing list_inactive member_09092025.xlsx')
CANONICAL_CSV = os.path.join(WORKING_DIR, 'canonical_member_list.csv')
OUTPUT_CSV = os.path.join(WORKING_DIR, 'canonical_member_list_cleaned.csv')
REMOVED_CSV = os.path.join(WORKING_DIR, 'removed_inactive_members.csv')


def read_inactive_emails():
    """Read inactive member emails from Excel file."""
    wb = load_workbook(INACTIVE_XLSX, read_only=True)
    ws = wb.active

    inactive_emails = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        email = row[0]
        if email:
            inactive_emails.add(email.strip().lower())

    wb.close()
    return inactive_emails


def remove_inactive_members():
    """Remove inactive members from canonical list."""

    print("Reading inactive member list...")
    inactive_emails = read_inactive_emails()
    print(f"  Found {len(inactive_emails)} inactive emails")

    print("\nReading canonical member list...")
    with open(CANONICAL_CSV, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        all_members = list(reader)
    print(f"  Found {len(all_members)} members")

    # Separate active from inactive
    active_members = []
    removed_members = []

    for member in all_members:
        email = member.get('email', '').strip().lower()
        if email in inactive_emails:
            removed_members.append(member)
        else:
            active_members.append(member)

    print(f"\nResults:")
    print(f"  Removed (inactive): {len(removed_members)}")
    print(f"  Kept (active): {len(active_members)}")

    # Write cleaned canonical list
    print(f"\nWriting {OUTPUT_CSV}...")
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(active_members)

    # Write removed members for reference
    print(f"Writing {REMOVED_CSV}...")
    with open(REMOVED_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(removed_members)

    print("\nDone!")

    return {
        'original_count': len(all_members),
        'removed_count': len(removed_members),
        'final_count': len(active_members),
    }


if __name__ == '__main__':
    stats = remove_inactive_members()
