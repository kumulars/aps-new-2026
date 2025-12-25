#!/usr/bin/env python3
"""
Merge MemberPress CSV with APS 2025 Symposium attendees to create canonical member list.

This script:
1. Reads the MemberPress CSV (2,103 members)
2. Reads the APS 2025 Symposium Excel (618 attendees)
3. Adds symposia_attended column
4. Merges 456 new attendees not in MemberPress
5. Updates 155 overlapping members to include "2025" in symposia_attended
6. Outputs canonical_member_list.csv
"""

import csv
import os
from openpyxl import load_workbook

# File paths
WORKING_DIR = os.path.dirname(os.path.abspath(__file__))
MEMBERPRESS_CSV = os.path.join(WORKING_DIR, 'memberpress_12_24_25.csv')
SYMPOSIUM_XLSX = os.path.join(WORKING_DIR, 'APS 2025 Membership Report.xlsx')
OUTPUT_CSV = os.path.join(WORKING_DIR, 'canonical_member_list.csv')

# Mapping symposium membership types to mepr_affiliation_type
MEMBERSHIP_TYPE_MAP = {
    'Graduate Student': 'academic',
    'Industry Professional': 'industry',
    'Non-Member': 'other',
    'PI/Faculty': 'academic',
    'Postdoc/Retired': 'academic',  # Default to academic
    'Undergraduate': 'academic',
    'Start-Up/PUI': 'academic',  # Primarily Undergraduate Institutions
    'Other': 'other',
}


def read_memberpress_csv():
    """Read MemberPress CSV and return dict keyed by lowercase email."""
    members = {}
    with open(MEMBERPRESS_CSV, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        for row in reader:
            email = row.get('email', '').strip().lower()
            if email:
                members[email] = row
    return members, fieldnames


def read_symposium_xlsx():
    """Read symposium Excel file and return list of attendees."""
    attendees = []
    wb = load_workbook(SYMPOSIUM_XLSX, read_only=True)
    ws = wb.active

    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value if cell.value else '')

    # Read data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Has data
            attendee = dict(zip(headers, row))
            attendees.append(attendee)

    wb.close()
    return attendees, headers


def create_canonical_list():
    """Create the canonical member list with symposia_attended field."""

    print("Reading MemberPress CSV...")
    members, mp_fieldnames = read_memberpress_csv()
    print(f"  Found {len(members)} members")

    print("\nReading APS 2025 Symposium Excel...")
    attendees, symp_headers = read_symposium_xlsx()
    print(f"  Found {len(attendees)} attendees")

    # Add symposia_attended to fieldnames
    output_fieldnames = list(mp_fieldnames) + ['symposia_attended', 'symposium_2025_membership_type']

    # Track statistics
    overlap_count = 0
    new_count = 0

    # Process symposium attendees
    symposium_emails = set()
    new_attendees = []

    for attendee in attendees:
        email = attendee.get('Email', '').strip().lower()
        if not email:
            continue

        symposium_emails.add(email)

        # Determine membership type from boolean flags
        membership_type = 'Other'
        if attendee.get('Graduate Student') == 'Yes':
            membership_type = 'Graduate Student'
        elif attendee.get('Industry Professional') == 'Yes':
            membership_type = 'Industry Professional'
        elif attendee.get('Academic Professional') == 'Yes':
            membership_type = 'PI/Faculty'
        elif attendee.get('Postdoc/Retired') == 'Yes':
            membership_type = 'Postdoc/Retired'
        elif attendee.get('Undergraduate Student') == 'Yes':
            membership_type = 'Undergraduate'
        elif attendee.get('Start-Up/PUI') == 'Yes':
            membership_type = 'Start-Up/PUI'

        if email in members:
            # Existing member attended symposium
            overlap_count += 1
            members[email]['symposia_attended'] = '2025'
            members[email]['symposium_2025_membership_type'] = membership_type
        else:
            # New attendee
            new_count += 1

            # Create new member record with mapped fields
            new_member = {field: '' for field in output_fieldnames}
            new_member['email'] = attendee.get('Email', '').strip()
            new_member['mepr_first_name'] = attendee.get('First', '') or ''
            new_member['mepr_last_name'] = attendee.get('Last', '') or ''
            new_member['mepr_professional_affiliation'] = attendee.get('Affiliation', '') or ''
            new_member['mepr_address_1'] = attendee.get('Address 1', '') or ''
            new_member['mepr_address_2'] = attendee.get('Address 2', '') or ''
            new_member['mepr_city'] = attendee.get('City', '') or ''
            new_member['mepr_state_or_province'] = attendee.get('State', '') or ''
            new_member['mepr_zip_or_postal_code'] = str(attendee.get('Postal Code', '') or '')
            new_member['mepr_country'] = attendee.get('Country', '') or ''
            new_member['mepr_phone'] = str(attendee.get('Phone', '') or '')
            new_member['mepr_title'] = attendee.get('Title', '') or ''

            # Map membership type to affiliation type
            affiliation = MEMBERSHIP_TYPE_MAP.get(membership_type, 'other')
            new_member['mepr_affiliation_type'] = affiliation

            # Set symposia attended
            new_member['symposia_attended'] = '2025'
            new_member['symposium_2025_membership_type'] = membership_type

            # Set status
            new_member['status'] = 'none'  # Match existing format

            new_attendees.append(new_member)

    # Members not at symposium get empty symposia_attended
    for email, member in members.items():
        if 'symposia_attended' not in member:
            member['symposia_attended'] = ''
            member['symposium_2025_membership_type'] = ''

    print(f"\nStatistics:")
    print(f"  Existing members who attended 2025: {overlap_count}")
    print(f"  New attendees to add: {new_count}")
    print(f"  Total canonical list: {len(members) + len(new_attendees)}")

    # Write output CSV
    print(f"\nWriting {OUTPUT_CSV}...")
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=output_fieldnames)
        writer.writeheader()

        # Write existing members first
        for member in members.values():
            writer.writerow(member)

        # Write new attendees
        for attendee in new_attendees:
            writer.writerow(attendee)

    print("Done!")
    print(f"\nOutput file: {OUTPUT_CSV}")

    return {
        'total_memberpress': len(members),
        'total_symposium': len(attendees),
        'overlap': overlap_count,
        'new_attendees': new_count,
        'total_canonical': len(members) + len(new_attendees),
    }


if __name__ == '__main__':
    stats = create_canonical_list()
