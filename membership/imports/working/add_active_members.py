#!/usr/bin/env python3
"""
Add missing active mailing list members to the canonical list.

This script:
1. Reads the active mailing list (2,162 emails)
2. Reads the cleaned canonical list (2,089 members)
3. Identifies members in active list but not in canonical
4. Adds them to create final canonical list
5. Verifies no duplicates
"""

import csv
import os
from openpyxl import load_workbook

WORKING_DIR = os.path.dirname(os.path.abspath(__file__))
ACTIVE_XLSX = os.path.join(WORKING_DIR, 'Mailing list_ACTIVE member_09092025.xlsx')
CANONICAL_CSV = os.path.join(WORKING_DIR, 'canonical_member_list_cleaned.csv')
OUTPUT_CSV = os.path.join(WORKING_DIR, 'canonical_member_list_final.csv')
ADDED_CSV = os.path.join(WORKING_DIR, 'added_active_members.csv')


def read_active_members():
    """Read active mailing list from Excel."""
    wb = load_workbook(ACTIVE_XLSX, read_only=True)
    ws = wb.active

    members = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        email = row[0]
        if email:
            members.append({
                'email': email.strip(),
                'email_lower': email.strip().lower(),
                'first_name': row[1] or '',
                'last_name': row[2] or '',
            })

    wb.close()
    return members


def add_missing_members():
    """Add missing active members to canonical list."""

    print("Reading active mailing list...")
    active_members = read_active_members()
    print(f"  Found {len(active_members)} members")

    print("\nReading canonical list...")
    with open(CANONICAL_CSV, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        canonical_rows = list(reader)

    canonical_emails = set(row.get('email', '').strip().lower() for row in canonical_rows)
    print(f"  Found {len(canonical_rows)} members")

    # Find members to add
    members_to_add = []
    for member in active_members:
        if member['email_lower'] not in canonical_emails:
            members_to_add.append(member)

    print(f"\nMembers to add: {len(members_to_add)}")

    # Create new rows for missing members
    new_rows = []
    for member in members_to_add:
        new_row = {field: '' for field in fieldnames}
        new_row['email'] = member['email']
        new_row['mepr_first_name'] = member['first_name']
        new_row['mepr_last_name'] = member['last_name']
        new_row['status'] = 'none'
        new_row['mepr_affiliation_type'] = ''  # Unknown
        new_rows.append(new_row)

    # Combine and deduplicate (keep first occurrence, which has more data)
    all_rows = canonical_rows + new_rows

    seen_emails = set()
    deduplicated_rows = []
    duplicates_removed = 0

    for row in all_rows:
        email = row.get('email', '').strip().lower()
        if email and email not in seen_emails:
            seen_emails.add(email)
            deduplicated_rows.append(row)
        elif email:
            duplicates_removed += 1

    all_rows = deduplicated_rows

    if duplicates_removed > 0:
        print(f"\nRemoved {duplicates_removed} duplicate emails (kept first occurrence with more data)")
    else:
        print(f"\nNo duplicates found")

    # Write final canonical list
    print(f"\nWriting {OUTPUT_CSV}...")
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)

    # Write added members for reference
    print(f"Writing {ADDED_CSV}...")
    with open(ADDED_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(new_rows)

    print("\nDone!")
    print(f"\nFinal canonical list: {len(all_rows)} members")

    return {
        'original_count': len(canonical_rows),
        'added_count': len(new_rows),
        'final_count': len(all_rows),
    }


if __name__ == '__main__':
    stats = add_missing_members()
